from dotenv import load_dotenv
import os
from binance.client import Client
import pandas as pd
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import threading

def fetch():
    tickers = client.get_ticker()
    crypto_data = []
    for ticker in tickers:
        if ticker['symbol'].endswith('USDT'):
            symbol = ticker['symbol'].replace('USDT', '')
            crypto_data.append({
                'Cryptocurrency Name': symbol,
                'Symbol': ticker['symbol'],
                'Current Price (USD)': float(ticker['lastPrice']),
                'Market Capitalization': float(ticker['quoteVolume']),
                '24h Trading Volume': float(ticker['volume']),
                'Price Change (24h %)': float(ticker['priceChangePercent']),
                'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
    return pd.DataFrame(crypto_data).sort_values('Market Capitalization', ascending=False).head(50)

def analyze(df):
    print("\n=== Crypto Market Analysis ===")
    top_5 = df.head(5)
    for idx, row in top_5.iterrows():
        print(f"{row['Cryptocurrency Name']}: ${row['Market Capitalization']:,.2f}")
    print(f"\nAverage Price of Top 50 Cryptocurrencies: ${df['Current Price (USD)'].mean():,.2f}")
    highest_gain = df.loc[df['Price Change (24h %)'].idxmax()]
    lowest_gain = df.loc[df['Price Change (24h %)'].idxmin()]
    print(f"\n24h Price Change Analysis:")
    print(f"Highest Gainer: {highest_gain['Cryptocurrency Name']} (+{highest_gain['Price Change (24h %)']}%)")
    print(f"Biggest Drop: {lowest_gain['Cryptocurrency Name']} ({lowest_gain['Price Change (24h %)']}%)")

def excel(filename='Top 50 Live Cryptocurrency Data'):
   
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, filename + '.xlsx')
    update_count = 0
    while True:
        try:
            df = fetch()
            analyze(df)
            if not os.path.exists(excel_path): Workbook().save(excel_path)
            
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name='Live Crypto Data', index=False)
                worksheet = writer.sheets['Live Crypto Data']
                
                for cell in worksheet[1]:
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

            update_count += 1
            print(f"Data update #{update_count} completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            time.sleep(300)
        except Exception as e:
            print(f"Error updating data: {str(e)}")
            time.sleep(60)

if __name__ == "__main__":
    load_dotenv()
    client = Client(os.environ['BINANCE_API_KEY'], os.environ['BINANCE_API_SECRET'], testnet=True)
    update_thread = threading.Thread(target=excel, daemon=True)
    update_thread.start()
    
    try:
        while True:
            print("Live update service is running... (Press Ctrl+C to stop)")
            time.sleep(60)
    except KeyboardInterrupt:
        print("\nStopping live updates. Please wait for any in-progress updates to complete...")